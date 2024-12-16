import os
import json
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, send_from_directory
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import chromedriver_autoinstaller
import re
import requests
import base64
from openpyxl import load_workbook
from openpyxl import Workbook
import zipfile
import PyPDF2
import pdfplumber
import io

app = Flask(__name__)

chromedriver_autoinstaller.install()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_PATH = os.path.join(BASE_DIR, "Promessa", "Promessa.xlsx")

if not os.path.exists(EXCEL_PATH):
    wb = load_workbook()
    ws = wb.active
    ws.title = "Promessas"
    ws.append(['CNPJ', 'Data Promessa'])
    wb.save(EXCEL_PATH)

SENHAS_ARQUIVO = os.path.join(BASE_DIR, 'templates', 'Senhas.txt')
DATA_JSON_ARQUIVO = os.path.join(BASE_DIR, 'templates', 'data.json')

app.secret_key = os.urandom(24)

cnpjs = []
data = []
nome = []
venc = []
saldo = []
emails = []


def desbloquear_pdf_em_memoria(caminho_pdf, senha):
    with open(caminho_pdf, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)

        if pdf_reader.is_encrypted:
            pdf_reader.decrypt(senha)

            pdf_writer = PyPDF2.PdfWriter()
            for pagina in range(len(pdf_reader.pages)):
                pdf_writer.add_page(pdf_reader.pages[pagina])

            pdf_bytes = io.BytesIO()
            pdf_writer.write(pdf_bytes)
            pdf_bytes.seek(0)

            return pdf_bytes
        else:
            return io.BytesIO(file.read())


def extrair_texto_boleto(caminho_pdf, senha):
    pdf_bytes = desbloquear_pdf_em_memoria(caminho_pdf, senha)

    texto_boleto = ""

    with pdfplumber.open(pdf_bytes) as pdf:
        for pagina in pdf.pages:
            texto_boleto += pagina.extract_text()

    return texto_boleto


def extrair_data_vencimento(texto):
    data_vencimento = re.search(r'\d{2}/\d{2}/\d{4}', texto)
    return data_vencimento.group(0) if data_vencimento else None


def extrair_valor(texto):
    valor = re.search(r'R\$\s?\d{1,3}(?:\.\d{3})*,\d{2}', texto)
    return valor.group(0) if valor else None


def extrair_numero_documento(texto):
    numero_documento = re.search(r'\d{2}\s?\d{5,10}', texto)
    if numero_documento:
        return numero_documento.group(0)
    return None


def processar_boletos_pasta(pasta, senha, cnpj, empresa, grouped_data):
    for arquivo in os.listdir(pasta):
        if arquivo.lower().endswith('.pdf') and arquivo.lower().startswith('boleto'):
            caminho_pdf = os.path.join(pasta, arquivo)

            try:
                texto = extrair_texto_boleto(caminho_pdf, senha)

                data_vencimento = extrair_data_vencimento(texto)
                valor = extrair_valor(texto)
                numero_documento = extrair_numero_documento(texto)

                def format_cnpj(cnpj):
                    cnpj = str(cnpj)
                    cnpj = re.sub(r'[^0-9]', '', cnpj)
                    return f'{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}'

                cnpj_formatados = format_cnpj(cnpjs)

                boleto_data = {
                    'cnpj': cnpj_formatados,
                    'empresa': empresa,
                    'boleto': arquivo,
                    'data_vencimento': data_vencimento,
                    'valor': valor,
                    'numero_documento': numero_documento
                }

                if empresa not in grouped_data:
                    grouped_data[empresa] = []
                grouped_data[empresa].append(boleto_data)

            except Exception as e:
                print(f"Erro ao processar {arquivo}: {e}")
                print("-" * 40)


def verificar_senha(token):
    with open(SENHAS_ARQUIVO, 'r') as f:
        senhas = f.readlines()

    for linha in senhas:
        if not linha.strip():
            continue

        dados = linha.strip().split(',')

        if len(dados) != 7:
            continue

        stored_token, cnpj, nome_empresa, data_geracao, vencimento_real, saldo_a_receber, email = dados
        if token == stored_token:
            session['cnpj'] = cnpj
            session['data'] = data_geracao
            session['nome_empresa'] = nome_empresa
            session['vencimento_real'] = vencimento_real
            session['saldo_a_receber'] = saldo_a_receber
            session['email'] = email
            cnpjs.append(cnpj)
            data.append(data_geracao)
            nome.append(nome_empresa)
            venc.append(vencimento_real)
            saldo.append(saldo_a_receber)
            emails.append(email)
            # senhas.remove(linha)
            with open(SENHAS_ARQUIVO, 'w') as f:
                f.writelines(senhas)
            return True
    return False


@app.route('/')
def index():
    return render_template('login.html')


@app.route('/verificar', methods=['POST'])
def verificar():
    if request.method == 'POST':
        token = request.form['senha']
        if verificar_senha(token):
            return redirect(url_for('loading'))
        else:
            return jsonify({"error": "Token inválido!"}), 403
    return render_template('login.html')


@app.route('/loading', methods=['GET'])
def loading():
    return render_template('loading.html')


@app.route('/process', methods=['GET'])
def process():
    with open(DATA_JSON_ARQUIVO, 'w') as json_file:
        json_file.write('')

    if cnpjs:
        def obter_token():

            url_token = "http://172.50.30.239:8962/REST/api/oauth2/v1/token"
            payload_token = {
                'grant_type': 'password',
                'username': 'INTEGRACAO',
                'password': 'INTEGRACAO'
            }
            headers_token = {
                'Content-Type': 'application/x-www-form-urlencoded'
            }
            try:
                response_token = requests.post(url_token, data=payload_token, headers=headers_token)

                if response_token.status_code == 201:
                    token_data = response_token.json()
                    print(token_data)
                    return token_data.get('access_token')
                else:
                    print(f"Erro na autenticação: {response_token.status_code}, {response_token.text}")
                    return None
            except Exception as e:
                print("Erro na CHAMADA a API token:", e)

        def processar_cnpj(cnpj, access_token, empresa, grouped_data):
            url_chatbot = "http://172.50.30.239:8962/rest/CHATBOT"
            headers_chatbot = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            params_chatbot = {'CNPJ': cnpj}
            pasta_cnpj = f"C:\\Projetos\\PORTAL\\Boletos\\{cnpj}"
            try:
                response_chatbot = requests.post(url_chatbot, headers=headers_chatbot, json=params_chatbot)

                if response_chatbot.status_code == 200:
                    response_json = response_chatbot.json()
                    print(f"Resposta do Chatbot para o CNPJ {cnpj}: {response_json}")
                    if isinstance(response_json, list) and len(response_json) > 0:
                        boletos_data = response_json
                        if not os.path.exists(pasta_cnpj):
                            os.makedirs(pasta_cnpj)
                            print(f"Pasta criada para o CNPJ {cnpj} em {pasta_cnpj}")

                        for idx, boleto_data in enumerate(boletos_data):
                            boleto_base64 = boleto_data.get('BOLETOS    ', None)
                            danfe_base64 = boleto_data.get('DANFE      ', None)

                            if boleto_base64:
                                try:
                                    boleto_bytes = base64.b64decode(boleto_base64)

                                    arquivo_destino = os.path.join(pasta_cnpj, f"boleto_{cnpj}_{idx + 1}.pdf")

                                    with open(arquivo_destino, "wb") as f:
                                        f.write(boleto_bytes)

                                    print(f"Boleto {idx + 1} para o CNPJ {cnpj} salvo como {arquivo_destino}")
                                except Exception as e:
                                    print(f"Erro ao decodificar o boleto {idx + 1} para o CNPJ {cnpj}: {str(e)}")
                            elif danfe_base64:
                                try:
                                    danfe_bytes = base64.b64decode(danfe_base64)

                                    arquivo_destino = os.path.join(pasta_cnpj, f"danfe_{cnpj}_{idx}.pdf")

                                    with open(arquivo_destino, "wb") as f:
                                        f.write(danfe_bytes)

                                    print(f"DANFE {idx} para o CNPJ {cnpj} salvo como {arquivo_destino}")
                                except Exception as e:
                                    print(f"Erro ao decodificar o DANFE {id} para o CNPJ {cnpj}: {str(e)}")
                            else:
                                print(f"Erro: Nenhum boleto ou DANFE encontrado para o CNPJ {cnpj}, item {idx + 1}.")
                        try:
                            if os.path.exists(pasta_cnpj):
                                processar_boletos_pasta(pasta_cnpj, senha='boletos', cnpj=cnpj, empresa=empresa,
                                                        grouped_data=grouped_data)
                        except Exception as e:
                            print("Erro para ler boletos:", e)
                    else:
                        print(f"Erro: Boleto não encontrado para o CNPJ {cnpj}.")
                elif response_chatbot.status_code == 500:
                    print(f"Erro 500 ao processar o CNPJ {cnpj}. Tentando o próximo.")
                    try:
                        if os.path.exists(pasta_cnpj):
                            processar_boletos_pasta(pasta_cnpj, senha='boletos', cnpj=cnpj, empresa=empresa,
                                                    grouped_data=grouped_data)
                    except Exception as e:
                        print("Erro para ler boletos:", e)
                else:
                    print(f"Erro ao processar o CNPJ {cnpj}: {response_chatbot.status_code}, {response_chatbot.text}")
                    try:
                        if os.path.exists(pasta_cnpj):
                            processar_boletos_pasta(pasta_cnpj, senha='boletos', cnpj=cnpj, empresa=empresa,
                                                    grouped_data=grouped_data)
                    except Exception as e:
                        print("Erro para ler boletos:", e)
            except Exception as e:
                print("Erro na CHAMADA a API boleto:", e)
                try:
                    if os.path.exists(pasta_cnpj):
                        processar_boletos_pasta(pasta_cnpj, senha='boletos', cnpj=cnpj, empresa=empresa,
                                                grouped_data=grouped_data)
                except Exception as e:
                    print("Erro para ler boletos:", e)

        def main(grouped_data):
            access_token = obter_token()
            # if not access_token:
            # return

            empresa = nome[0]
            cnpj = cnpjs[0]

            print(f"Iniciando o processamento para a empresa {empresa}, CNPJ {cnpj}...")

            processar_cnpj(cnpj, access_token, empresa, grouped_data)

            if grouped_data:
                datahora = datetime.now().strftime("%d/%m/%Y %H:%M")

                return (grouped_data)

            print("Processamento completo.")

        empresa = nome[0]

        grouped_data = {}

        all_grouped_data = {}

        data = []

        main(grouped_data)

        def format_cnpj(cnpj):
            cnpj = str(cnpj)
            cnpj = re.sub(r'[^0-9]', '', cnpj)
            return f'{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}'

        cnpj_formatados = format_cnpj(cnpjs)

        data.append({
            'cnpj': cnpj_formatados,
        })

        df = pd.DataFrame(data)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.expand_frame_repr', False)

        columns_to_check = ['cnpj']

        for col_name in columns_to_check:
            if col_name not in df.columns:
                df[col_name] = ''

        colunas_pd = ['cnpj']

        df_filtrado = df[colunas_pd]

        if not grouped_data:
            for registro in data:
                print("Registro:", registro)

                if empresa not in grouped_data:
                    grouped_data[empresa] = []
                grouped_data[empresa].append(registro)

        data = df_filtrado.to_dict(orient='records')

        datahora = datetime.now()
        datahora = datahora.strftime("%d/%m/%Y %H:%M")

        if grouped_data:
            if empresa not in all_grouped_data:
                all_grouped_data[empresa] = []
            all_grouped_data[empresa].append(grouped_data)

        with open(DATA_JSON_ARQUIVO, 'a') as json_file:
            json.dump(all_grouped_data, json_file)
        return jsonify({'empresa': empresa, 'datahora': datahora})


@app.route('/data', methods=['GET'])
def get_data():
    with open(DATA_JSON_ARQUIVO, 'r') as json_file:
        data = json.load(json_file)

    return jsonify(data)


@app.route('/empresas', methods=['GET'])
def empresas():
    empresa_nome = request.args.get('empresa')
    datahora = request.args.get('datahora')

    with open(DATA_JSON_ARQUIVO, 'r') as json_file:
        data = json.load(json_file)

    if empresa_nome in data:
        empresa_data = data[empresa_nome]
    else:
        empresa_data = []

    return render_template('empresas.html', empresa=empresa_nome, datahora=datahora, empresa_data=empresa_data)


@app.route('/gif/<int:gif_number>')
def get_gif(gif_number):
    gif_path = os.path.join(app.root_path, 'templates', 'static', 'images', f'Animation - {gif_number}.GIF')
    return send_file(gif_path, mimetype='image/gif')


@app.route('/gerar_boleto', methods=['POST'])
def gerar_boleto():
    cnpj = request.form.get('cnpj')
    data_promessa = request.form.get('data_promessa')
    nome_boleto = request.form.get('nome_boleto')
    nome_danfe = nome_boleto.replace('boleto', 'danfe')

    cnpj = cnpj.replace('.', '').replace('/', '').replace('-', '')
    pasta_cnpj = os.path.join(BASE_DIR, "Boletos", cnpj)

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Promessas"
        ws.append(['CNPJ', 'Data Promessa'])
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([cnpj, data_promessa])
    wb.save(EXCEL_PATH)

    with open(SENHAS_ARQUIVO, 'r', encoding='utf-8') as file:
        linhas_txt = file.readlines()

    total_linhas_txt = len(linhas_txt)

    if os.path.exists(EXCEL_PATH):
        df = pd.read_excel(EXCEL_PATH)

        cnpjs_unicos = df['CNPJ'].drop_duplicates()

        total_promessas = len(cnpjs_unicos)
    else:
        total_promessas = 0

    if total_linhas_txt > 0:
        porcentagem_promessas = (total_promessas / total_linhas_txt) * 100
    else:
        porcentagem_promessas = 0

    print(f"Total de Linhas no TXT: {total_linhas_txt}")
    print(f"Total de Promessas na Planilha: {total_promessas}")
    print(f"Porcentagem de Promessas: {porcentagem_promessas:.2f}%")

    resultado_excel = r'C:\Projetos\PORTAL\Promessa\resultado_analise.xlsx'
    wb_resultado = Workbook()
    ws_resultado = wb_resultado.active
    ws_resultado.title = "Análise de Promessas"

    ws_resultado.append(['Total Linhas TXT', 'Total Promessas', 'Porcentagem de Promessas'])


    return {
        'total_linhas_txt': total_linhas_txt,
        'total_promessas': total_promessas,
        'porcentagem_promessas': porcentagem_promessas
    }

    boleto_path = os.path.join(pasta_cnpj, nome_boleto)
    danfe_path = os.path.join(pasta_cnpj, nome_danfe)

    if os.path.exists(boleto_path) and boleto_path.endswith('.pdf') and os.path.exists(
            danfe_path) and danfe_path.endswith('.pdf'):
        try:
            print("Boleto e DANFE encontrados, enviando...")

            boleto_url = f'/download/{cnpj}/{nome_boleto}'
            danfe_url = f'/download/{cnpj}/{nome_danfe}'

            return jsonify({
                'boleto_url': boleto_url,
                'danfe_url': danfe_url
            })
        except Exception as e:
            print(f"Erro ao enviar arquivos: {e}")
            return jsonify({'erro': 'Erro ao enviar os arquivos'}), 500

    else:
        print(f"Arquivos não encontrados: {boleto_path} ou {danfe_path}")
        return jsonify({
            'erro': 'Boleto ou DANFE não encontrado. Por favor, entre em contato com o chatbot para obter os arquivos.'
        }), 404


@app.route('/download/<cnpj>/<file_name>')
def download_file(cnpj, file_name):
    file_path = os.path.join(BASE_DIR, "Boletos", cnpj, file_name)
    if os.path.exists(file_path):
        return send_from_directory(os.path.dirname(file_path), file_name, as_attachment=True)
    else:
        return jsonify({'erro': 'Arquivo não encontrado'}), 404



if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
